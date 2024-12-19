import json
import pickle
import time

import pytest
from openpyxl import load_workbook  # Importing openpyxl for reading Excel files
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By


# Page Object Model for Login Page
class LoginPage:
    def __init__(self, driver):
        self.driver = driver
        self.username_input = (By.ID, "username")
        self.password_input = (By.NAME, "password")
        self.login_button = (By.XPATH, "//button[normalize-space()='Sign In']")
        self.error_message = (By.XPATH, "//div[@class='message alert alert-danger']")

    def enter_username(self, username):
        self.driver.find_element(*self.username_input).send_keys(username)

    def enter_password(self, password):
        self.driver.find_element(*self.password_input).send_keys(password)

    def click_login(self):
        self.driver.find_element(*self.login_button).click()

    def get_error_message(self):
        try:
            return self.driver.find_element(*self.error_message).text
        except NoSuchElementException:
            return None

    def is_login_successful(self):
        return "EMS : Administer/Dashboard" in self.driver.title


# Utility function to read Excel data
def read_excel_data(file_name):
    workbook = load_workbook(file_name)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        username, password = row
        data.append((username, password))
    return data


# Pytest fixture to set up and tear down the WebDriver
@pytest.fixture(params=["chrome", "firefox"])
def driver(request):
    browser = request.param
    if browser == "chrome":
        service = Service('./drivers/chromedriver-win64/chromedriver.exe')
        driver = webdriver.Chrome(service=service)
    elif browser == "firefox":
        service = Service('./drivers/geckodriver-v0.35.0-win64/geckodriver.exe')
        driver = webdriver.Firefox(service=service)
    yield driver
    driver.quit()


# Parametrize test cases based on Excel data
@pytest.mark.parametrize("username, password", read_excel_data('test_data.xlsx'))
def test_login(driver, username, password):
    driver.get("https://ems-test.amaderit.net/")
    driver.maximize_window()

    # Create an instance of the LoginPage
    login_page = LoginPage(driver)

    # Perform login actions
    login_page.enter_username(username)
    login_page.enter_password(password)
    login_page.click_login()

    time.sleep(5)

    # Check for login error or success
    error_message = login_page.get_error_message()
    if error_message == "Please enter your correct username and password":
        raise AssertionError("Login failed, username and/or password is invalid")

    # Assert successful login
    assert login_page.is_login_successful(), "Dashboard not found"

    # Save session and local storage data after successful login
    local_storage = driver.execute_script("return window.localStorage;")
    with open("local_storage.pkl", "wb") as f:
        pickle.dump(local_storage, f)

    # Optionally capture cookies
    # cookies = driver.get_cookies()
    # with open("cookies.pkl", "wb") as file:
    #     pickle.dump(cookies, file)
