import json
import pickle

import pytest
from openpyxl import load_workbook  # Importing openpyxl for reading Excel files
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, WebDriverException
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import time  # Import the time module

from selenium.webdriver.support.wait import WebDriverWait

# Read Excel data
def read_excel_data(file_name):
    workbook = load_workbook(file_name)
    sheet = workbook.active  # Get the first sheet
    data = []

    # Loop through the rows in the Excel file and collect the username and password
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
        username, password = row
        data.append((username, password))
    return data

# Pytest fixture to handle WebDriver setup and teardown
#@pytest.fixture(scope="session")
# Pytest fixture to handle browser setup
@pytest.fixture(params=["chrome", "firefox"], scope="session")
def driver(request):
    #browser = request.config.getoption("--driver")  # Get the browser from the command-line option
    browser = request.param
    if browser == "chrome":
        service = Service('./drivers/chromedriver-win64/chromedriver.exe')
        driver = webdriver.Chrome(service=service)
    elif browser == "firefox":
        service = Service('./drivers/geckodriver-v0.35.0-win64/geckodriver.exe')
        driver = webdriver.Firefox(service=service)
    # service = Service('./drivers/chromedriver-win64/chromedriver.exe')
    # driver = webdriver.Chrome(service=service)
    #driver.implicitly_wait(10)
    yield driver  # This returns the driver to the test

    # Teardown: Close the browser (this is the teardown)
    driver.quit() # Ensures that the browser is closed after the test

@pytest.mark.parametrize("username, password", read_excel_data('test_data.xlsx'))
def test_login(driver, username, password):
    print("Starting the test.")

    # Open EMS
    driver.get("https://ems-test.amaderit.net/")
    # driver.get("https://www.google.com/")
    # Maximize the browser window
    driver.maximize_window()
    print("Navigated to EMS page.")
    print("Page title is:", driver.title)

    try:
        # Locate the username and password fields and input login credentials
        print("Trying to locate username and password fields.")
        username_field = driver.find_element(By.ID, "username")
        password_field = driver.find_element(By.NAME, "password")

    except NoSuchElementException:
        raise AssertionError("Test failed: Username and password fields are not available.")

    # Enter the login credentials
    print("Entering the username and password.")
    # username_field.send_keys("adming1")
    # password_field.send_keys("12345678")

    username_field.send_keys(username)
    password_field.send_keys(password)

    # Locate the login button and click it
    login_button = driver.find_element(By.XPATH, "//button[normalize-space()='Sign In']")
    login_button.click()
    time.sleep(5)
    # Check for a failed login indicator
    try:
        error_message = driver.find_element(By.XPATH, "//div[@class='message alert alert-danger']")
        if error_message.text== "Please enter your correct username and password":
           raise AssertionError("Login failed, username and/or password is invalid")

    except NoSuchElementException:
        # Wait for some time to observe the result
        time.sleep(5)

        browser_title = driver.title
        url = driver.current_url
        # success_message = driver.find_element(By.XPATH, "//div[@class='message alert alert-success']")

        # Print the page title to the console
        print("Page title is after login:", browser_title)

        # Save the session cookies to a file after login
        # cookies = driver.get_cookies()
        # print(f"Cookies after login: {cookies}")
        # with open("cookies.pkl", "wb") as file:
        #     pickle.dump(cookies, file)

        # Capture local storage data after login
        local_storage = driver.execute_script("return window.localStorage;")

        # Set local storage data before running the next test
        for key, value in local_storage.items():
            value_json = json.dumps(value)  # Properly escapes the value
            driver.execute_script(f"window.localStorage.setItem('{key}', {value_json});")

        # Save local storage data to a file
        with open("local_storage.pkl", "wb") as f:
            pickle.dump(local_storage, f)

        # Wait for user input to keep the browser open
        # input("Press Enter to continue...")

        # Assert that login was successful by checking the title, URL, or page content
        # assert "administersa" in url.lower(), "administer is not found in url"
        # assert "Logged in successfullyyyy" in success_message.text, "Dashboard not found"
        assert "EMS : Administer/Dashboard" in browser_title, "Dashboard not found"